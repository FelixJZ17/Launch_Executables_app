import openpyxl
import subprocess
import time
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import psutil

def lanzar_programas():
    # 1. Seleccionar el mismo archivo Excel
    file_path = obtener_archivo_excel()

    if not file_path:
        return

    try:
        # 2. Cargar el libro de Excel
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active # Lee la hoja activa

        # 3. Recorrer filas empezando desde la fila 2 (min_row=2)
        # Columna 1 (A): Ruta | Columna 2 (B): Ejecutable
        lanzados = 0
        omitidos = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ruta = str(row[0]).strip() if row[0] else None
            ejecutable = str(row[1]).strip() if row[1] else None

            # Si la fila está vacía, saltar
            if not ruta or not ejecutable or ruta == "None":
                continue

            full_path = os.path.join(ruta, ejecutable)

            # --- NUEVA LÓGICA DE DETECCIÓN ---
            if esta_ejecutandose(ejecutable):
                print(f"Omitido: {ejecutable} ya está en ejecución.")
                omitidos += 1
                continue
            # ---------------------------------
            
            if os.path.exists(full_path) and not esta_ejecutandose(ejecutable):
                print(f"Iniciando: {ejecutable}")
                subprocess.Popen([full_path], cwd=ruta, shell=True)
                lanzados += 1
                time.sleep(3) # Espera de 3 segundos entre programas
            else:
                print(f"No se encontró: {full_path}")

        messagebox.showinfo("Finalizado", 
                            f"Procesados:\n- Lanzados: {lanzados}\n- Ya abiertos (omitidos): {omitidos}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al leer el Excel: {e}")

def esta_ejecutandose(nombre_ejecutable):
    """
    Revisa si existe un proceso activo con el nombre del ejecutable.
    """
    for proc in psutil.process_iter(['name']):
        try:
            # Comparamos nombres en minúsculas para evitar errores de coincidencia
            if proc.info['name'].lower() == nombre_ejecutable.lower():
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False

def cerrar_programas():
    # 1. Seleccionar el mismo archivo Excel
    file_path = obtener_archivo_excel()

    if not file_path:
        return

    try:
        # 2. Cargar el libro de Excel
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        cerrados = 0
        errores = 0

        # 3. Recorrer la lista
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # En el cierre solo nos importa el nombre del ejecutable (columna B)
            ejecutable = str(row[1]).strip() if row[1] else None

            if not ejecutable or ejecutable == "None":
                continue

            encontrado = False
            # 4. Buscar procesos activos con ese nombre
            for proc in psutil.process_iter(['name']):
                try:
                    if proc.info['name'].lower() == ejecutable.lower():
                        proc.terminate()  # Intenta cerrar elegantemente
                        # proc.kill() # Usa esto si el programa se resiste a cerrar
                        encontrado = True
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    
                    continue
            
            if encontrado:
                print(f"Cerrando: {ejecutable}")
                cerrados += 1
            else:
                print(f"No se encontró {ejecutable} en ejecución.")
                errores += 1

        messagebox.showinfo("Proceso Terminado", 
                            f"Se han cerrado instancias de {cerrados} programas de la lista. {errores} no se han encontrado para cerrar o no estaban abiertos.")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al cerrar programas: {e}")

def obtener_archivo_excel():
    """Abre el explorador de archivos para seleccionar el Excel."""
    file_path = filedialog.askopenfilename(
        title="Selecciona el Excel de configuración",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )
    return file_path

def main():
    root = tk.Tk()
    root.title("Gestor de Programas")
    root.geometry("350x250")
    root.eval('tk::PlaceWindow . center') # Centrar ventana

    # Estilo y Padding
    main_frame = tk.Frame(root, padx=20, pady=20)
    main_frame.pack(expand=True)

    tk.Label(main_frame, text="¿Qué acción desea realizar?", font=("Arial", 12, "bold")).pack(pady=10)

    # Botón Abrir
    btn_abrir = tk.Button(main_frame, text="ABRIR PROGRAMAS", width=25, command=lanzar_programas)
    btn_abrir.pack(pady=5)

    # Botón Cerrar
    btn_cerrar = tk.Button(main_frame, text="CERRAR PROGRAMAS", width=25, command=cerrar_programas)
    btn_cerrar.pack(pady=5)

    # Botón Salir
    btn_salir = tk.Button(main_frame, text="Salir", width=10, command=root.destroy)
    btn_salir.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()